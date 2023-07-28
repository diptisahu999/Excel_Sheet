from django.shortcuts import render
from GreenBuildingManagement.models import BmsGreenBuildingManageDevice,BmsGreenBuildingData
from GreenBuildingManagement.serializers import *
from rest_framework.response import Response 
from rest_framework.decorators import api_view
from rest_framework import status
# Create your views here.
import requests
import time
import json

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from django.http import HttpResponse



# Bms_Module crud

token = []
@api_view(['GET', 'POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def Green_list(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingManageDevice.objects.all()        
        bms_module_serializer = BmsGreenBuildingManageDeviceSerializer(bms_module, many=True)
        return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Lists", "response":bms_module_serializer.data},status=status.HTTP_200_OK)
    
 
    elif request.method == 'POST':
        module_serializer = BmsGreenBuildingManageDeviceSerializer(data=request.data)
        if module_serializer.is_valid():
            module_serializer.save()
            return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Added Sucessfuly!!","response":module_serializer.data},status=status.HTTP_201_CREATED) 
        return Response({"status_code":401,"responce":module_serializer.errors},status=status.HTTP_400_BAD_REQUEST) 
    
    elif request.method == 'DELETE':
        count = BmsGreenBuildingManageDevice.objects.all().delete()
        return Response({'message': '{} Green Building Manage Device was deleted successfully!'.format(count[0])}, status=status.HTTP_204_NO_CONTENT)
    
    
### GET Airveda All data
    
@api_view(['GET', 'POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def Green_data_list(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingData.objects.filter(manage_device=1)       
        bms_module_serializer = BmsGreenBuildingManageDataSerializer(bms_module, many=True)
        return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Lists", "response":bms_module_serializer.data},status=status.HTTP_200_OK)
    
    elif request.method == 'POST':
        module_serializer = BmsGreenBuildingManageDataSerializer(data=request.data)
        if module_serializer.is_valid():
            module_serializer.save()
            return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Added Sucessfuly!!","response":module_serializer.data},status=status.HTTP_201_CREATED) 
        return Response({"status_code":401,"responce":module_serializer.errors},status=status.HTTP_400_BAD_REQUEST) 
    
 
 ## GET Aura All Data   
    
@api_view(['GET', 'POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def Green_data_list_Aura(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingData.objects.filter(manage_device=2)       
        bms_module_serializer = BmsGreenBuildingManageDataSerializer(bms_module, many=True)
        return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Lists", "response":bms_module_serializer.data},status=status.HTTP_200_OK)
    


## last data get from the table Airveda

@api_view(['GET', 'POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def Green_last_data_list(request):
    if request.method == 'GET':
        # Retrieve the last data entry from the BmsGreenBuildingDatass table
        try:
            # bms_module = BmsGreenBuildingDatass.objects.latest('id')
            bms_module = BmsGreenBuildingData.objects.filter(manage_device=1).last()
        except BmsGreenBuildingDatass.DoesNotExist:
            return Response({"data": "false", "status_code": 404, "message": "No data found"}, status=status.HTTP_404_NOT_FOUND)
        
        # Serialize the data entry
        # bms_module_serializer = BmsGreenBuildingManageDataSerializersss(bms_module)
        bms_module_serializer = BmsGreenBuildingManageDataSerializer(bms_module)
        
        
        return Response({"data": "true", "status_code": 200, "message": "Green Building Manage Device Lists", "response": bms_module_serializer.data}, status=status.HTTP_200_OK)


## last data get from the table Airveda


@api_view(['GET', 'POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def Green_last_data_list_aura(request):
    if request.method == 'GET':
        # Retrieve the last data entry from the BmsGreenBuildingDatass table
        try:
            # bms_module = BmsGreenBuildingDatass.objects.latest('id')
            bms_module = BmsGreenBuildingData.objects.filter(manage_device=2).last()
        except BmsGreenBuildingDatass.DoesNotExist:
            return Response({"data": "false", "status_code": 404, "message": "No data found"}, status=status.HTTP_404_NOT_FOUND)
        
        # Serialize the data entry
        # bms_module_serializer = BmsGreenBuildingManageDataSerializersss(bms_module)
        bms_module_serializer = BmsGreenBuildingManageDataSerializer(bms_module)
        
        
        return Response({"data": "true", "status_code": 200, "message": "Green Building Manage Device Lists", "response": bms_module_serializer.data}, status=status.HTTP_200_OK)





### Excel Sheet in Airveda


@api_view(['GET'])
def export_as_excel_airveda(request):
    data1={
        "from_date": "16:11 26/07/23"
    }
    # data1 = request.data  # Get the data from the POST request body
    # print(data1)

    # data = BmsGreenBuildingData.objects.filter(created_at=data1['from_date'])
    # data = BmsGreenBuildingData.objects.all()
    data = BmsGreenBuildingData.objects.filter(manage_device=1)
    
    # print(data)


    serializer = BmsGreenBuildingManageDataSerializer(data, many=True)
    serialized_data = serializer.data
    # print(serialized_data)

    #
    workbook = Workbook()
    sheet = workbook.active

    # Write the column headers
    headers = list(serialized_data[0].keys()) if serialized_data else []  # Get the keys from the first item in data
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Write the data rows
    for row_num, row_data in enumerate(serialized_data, 2):
        for col_num, field_name in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell_value = str(row_data.get(field_name, ""))  # Convert the value to a string
            sheet[f"{col_letter}{row_num}"] = cell_value

    # Create a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response


### Excel Sheet in Aura


@api_view(['GET'])
def export_as_excel_aura(request):
    data1={
        "from_date": "16:11 26/07/23"
    }
    # data1 = request.data  # Get the data from the POST request body
    # print(data1)

    # data = BmsGreenBuildingData.objects.filter(created_at=data1['from_date'])
    # data = BmsGreenBuildingData.objects.all()
    data = BmsGreenBuildingData.objects.filter(manage_device=2)
    
    # print(data)


    serializer = BmsGreenBuildingManageDataSerializer(data, many=True)
    serialized_data = serializer.data
    # print(serialized_data)

    #
    workbook = Workbook()
    sheet = workbook.active

    # Write the column headers
    headers = list(serialized_data[0].keys()) if serialized_data else []  # Get the keys from the first item in data
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Write the data rows
    for row_num, row_data in enumerate(serialized_data, 2):
        for col_num, field_name in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell_value = str(row_data.get(field_name, ""))  # Convert the value to a string
            sheet[f"{col_letter}{row_num}"] = cell_value

    # Create a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response
    
    
    
# @api_view(['POST'])
# def export_as_excel(request):
#     if request.method == 'POST':
#         # data1={
#         #     "from_date": "16:11 26/07/23"
#         # }
#         data1 = request.data  # Get the data from the POST request body
#         # print(data1)
            
#         data = BmsGreenBuildingDatass.objects.filter(created_at=data1['from_date'])
#         data = BmsGreenBuildingDatass.objects.all()
#         # print(data)

    
#         serializer = BmsGreenBuildingManageDataSerializersss(data, many=True)
#         serialized_data = serializer.data
#         # print(serialized_data)

#         #
#         workbook = Workbook()
#         sheet = workbook.active

#         # Write the column headers
#         headers = list(serialized_data[0].keys()) if serialized_data else []  # Get the keys from the first item in data
#         for col_num, header in enumerate(headers, 1):
#             col_letter = get_column_letter(col_num)
#             sheet[f"{col_letter}1"] = header

#         # Write the data rows
#         for row_num, row_data in enumerate(serialized_data, 2):
#             for col_num, field_name in enumerate(headers, 1):
#                 col_letter = get_column_letter(col_num)
#                 cell_value = str(row_data.get(field_name, ""))  # Convert the value to a string
#                 sheet[f"{col_letter}{row_num}"] = cell_value

#         # Create a response object
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

#         # Save the workbook to the response
#         workbook.save(response)
#         return response
        

#         return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Added Sucessfuly!!"},status=status.HTTP_201_CREATED)
#     # return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Added Sucessfuly!!","response":module_serializer.data},status=status.HTTP_201_CREATED) 
# # return Response({"status_code":401,"responce":module_serializer.errors},status=status.HTTP_400_BAD_REQUEST)


    
    
 
# @api_view(['GET', 'PUT', 'DELETE'])
# # @permission_classes([IsAuthenticated])
# def green_detail(request, pk):
#     try: 
#         bms_module = BmsGreenBuildingManageDevice.objects.get(pk=pk) 
#     except BmsGreenBuildingManageDevice.DoesNotExist: 
#         return Response({'message': 'The Green Building Manage Device does not exist'}, status=status.HTTP_404_NOT_FOUND) 
         
 
#     if request.method == 'GET': 
#         module_serializer = BmsGreenBuildingManageDeviceSerializer(bms_module) 
#         return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device Get Successfully", "response":module_serializer.data},status=status.HTTP_200_OK)  
 
#     elif request.method == 'PUT': 
#         module_serializer = BmsGreenBuildingManageDeviceSerializer(bms_module, data=request.data) 
#         if module_serializer.is_valid(): 
#             module_serializer.save() 
#             return Response({"data":"true","status_code": 200, "message": "Green Building Manage Device updated Sucessfuly!!","response":module_serializer.data},status=status.HTTP_201_CREATED) 
#         return Response({"status_code":401,"responce":module_serializer.errors},status=status.HTTP_400_BAD_REQUEST)  
         
 
#     elif request.method == 'DELETE': 
#         bms_module.delete() 
#         return Response({'message': 'Green Building Manage Device was deleted successfully!'}, status=status.HTTP_204_NO_CONTENT)
    
    
@api_view(['GET', 'POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def GetToken(request):
    if request.method == 'GET':
        token.clear()
        bms_module = BmsGreenBuildingManageDevice.objects.all()        
        # print(bms_module.values())
        data_user= bms_module.values()
        a = data_user[0]['user_name']
        # print(a)
       
        base_api = data_user[0]['api_key']
        data = {
        "email": data_user[0]['user_name'],
        "password": data_user[0]['password']
        }
        
        response = requests.post(str(base_api),json=data)
        data = response.json()
        token.append(data)
        # My_data = {"idToken":data['idToken']}
        # {"idToken":data['idToken']}
        
        return Response({"data":"true","status_code": 200, "message": "Get Token Lists", "response":data} ,status=status.HTTP_200_OK)
    
    

import requests
from datetime import datetime, timedelta

@api_view(['GET', 'POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def refrece_token(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingManageDevice.objects.all()        
        print(bms_module.values())
        data_user = bms_module.values()
        a = data_user[0]['user_name']
        # print(a)
        data_user[0]['password']
        base_api = data_user[0]['api_key']

        expiration_time = datetime.now() + timedelta(seconds=15)  # Set token expiration time to 3 minutes

        data = {
            "email": data_user[0]['user_name'],
            "password": data_user[0]['password'],
            "expiration_time": expiration_time.strftime('%Y-%m-%d %H:%M:%S')
        }
        
        response = requests.post(str(base_api), json=data)
        data = response.json()
        # My_data = {"refreshToken":data['refreshToken']}
        return Response({
            "data": "true",
            "status_code": 200,
            "message": "Get Refresh Lists",
            "response": data
        }, status=status.HTTP_200_OK)



@api_view(['GET','POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def lastData(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingManageDevice.objects.all()        
        # print(bms_module.values())
        data_user= bms_module.values()
        a = data_user[0]['model_no']
        # print(token[0]['idToken'])
        # data = request.data
        # idToken= data['idToken']
        # token = ""
        last_token = token[0]['idToken']
        
        # device_id = data['device_id']
        base_api = 'https://dashboard.airveda.com/api/data/last-hour/'
    
        
        headers = {
            'Authorization': f'Bearer {last_token}',
            'Content-Type': 'application/json'
        }
        data = {
            "deviceId":a
        }

        response = requests.post(base_api, headers=headers, json=data)
        print(response)
        device_data = response.json()
        print(device_data)
        return Response({
            "data": "true",
            "status_code": 200,
            "message": "Get Device data Lists",
            "response": device_data
        }, status=status.HTTP_200_OK)
        
        
        

# @api_view(['GET','POST', 'DELETE'])
# # @permission_classes([IsAuthenticated])
# def lastDeviceData(request):
#     if request.method == 'GET':
#         bms_module = BmsGreenBuildingManageDevice.objects.all()        
#         # print(bms_module.values())
#         data_user= bms_module.values()
#         a = data_user[0]['model_no']
#         b = data_user[0]['id']
#         print(b)

#         last_token = token[0]['idToken']
        
#         # device_id = data['device_id']
#         base_api = 'https://dashboard.airveda.com/api/data/last-hour/'
    
        
#         headers = {
#             'Authorization': f'Bearer {last_token}',
#             'Content-Type': 'application/json'
#         }
#         data = {
#             "deviceId":a
#         }

#         response = requests.post(base_api, headers=headers, json=data)
#         device_data = response.json()
#         print(device_data) 
#         leed_data = []
#         leed_data.clear()
#         for item in device_data['data']: 
#             leed_data1 = {
#                     "manage_device":b,
#                     "leed_data": {
#                         "temperature": item["temperature"],
#                         "pm10": item["pm10"],
#                         "voc": item["voc"],
#                         "humidity": item["humidity"],
#                         "aqi": item["aqi"],
#                         "pm25": item["pm25"],
#                         "viral_index": item["viral_index"]
#                     }
#                 }
#             print(leed_data1)
#             module_serializer = BmsGreenBuildingManageDataSerializersss(data=leed_data1)
#             # print(module_serializer)
#             if module_serializer.is_valid():
#                 module_serializer.save()
#                 # print("i anm pass")
#                 leed_data.append(module_serializer)
#             else:
#                 # Handle invalid data, e.g., log the error or skip saving this data.
#                 print("Invalid data:", module_serializer.errors)
#                 leed_data.append(leed_data1)
            
#         # print(leed_data)
#         # module_serializer = BmsGreenBuildingManageDataSerializersss(data=leed_data)
#         # print(module_serializer)
#         # if module_serializer.is_valid():
#         #     module_serializer.save()
#         #     print("i anm pass")
#         # else:
#         #     # Handle invalid data, e.g., log the error or skip saving this data.
#         #     print("Invalid data:", module_serializer.errors)
            
        
#         # manage_device
#         # temperature
#         # pm10
#         # voc
#         # humidity
    
#         # aqi
#         # pm25
#         # viral_index
#         # print(selected_data)
#         # print(device_data)
        
#         # aqi = device_data['data']
#         # print(aqi)
        
#         # data={
                
#         #         "temperature":item["temperature"],
#         #         "pm10": item["pm10"],
#         #         "co2": 882,
#         #         "voc":  item["voc"],
#         #         "aqi":112 , 
#         #         "humidity": item["humidity"],
#         #         "battery": 5,
#         #         "pm25":111,
#         #         "viral_index": 82
#         #     }
#         # for i in selected_data:
#         #     module_serializer = BmsGreenBuildingManageDataSerializersss(data=i)   
#         #     module_serializer.is_valid()
#         #     module_serializer.save()
            

#         # module_serializer = BmsGreenBuildingManageDataSerializersss(data=selected_data)   
#         # module_serializer.is_valid()
#         # module_serializer.save()
#         # print(module_serializer)
#         return Response({
#             "data": "true",
#             "status_code": 200,
#             "message": "Get Device data Lists",
#             "response": module_serializer.data
#         }, status=status.HTTP_200_OK)
        
        
## excel Sheet parameter       


@api_view(['GET','POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def lastDeviceData(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingManageDevice.objects.all()        
        # print(bms_module.values())
        data_user= bms_module.values()
        a = data_user[0]['model_no']
        b = data_user[0]['id']
        # print(token[0]['idToken'])
        # data = request.data
        # idToken= data['idToken']
        # token = ""
        last_token = token[0]['idToken']
        
        # device_id = data['device_id']
        base_api = 'https://dashboard.airveda.com/api/data/last-hour/'
    
        
        headers = {
            'Authorization': f'Bearer {last_token}',
            'Content-Type': 'application/json'
        }
        data = {
            "deviceId":a
        }

        response = requests.post(base_api, headers=headers, json=data)
        device_data = response.json()
        print(device_data) 
        selected_data = []
        selected_data.clear()
        for item in device_data['data']:
            selected_item = {
                "manage_device":b,
                "temperature": item["temperature"],
                "pm10": item["pm10"],
                "voc": item["voc"],
                "humidity": item["humidity"],
                "aqi": item["aqi"],
                "pm25": item["pm25"],
                "viral_index": item["viral_index"]
            }
            selected_data.append(selected_item)
        
        # manage_device
        # temperature
        # pm10
        # voc
        # humidity
    
        # aqi
        # pm25
        # viral_index
        # print(selected_data)
        # print(device_data)
        
        # aqi = device_data['data']
        # print(aqi)
        
        # data={
                
        #         "temperature":item["temperature"],
        #         "pm10": item["pm10"],
        #         "co2": 882,
        #         "voc":  item["voc"],
        #         "aqi":112 , 
        #         "humidity": item["humidity"],
        #         "battery": 5,
        #         "pm25":111,
        #         "viral_index": 82
        #     }
        for i in selected_data:
            module_serializer = BmsGreenBuildingManageDataSerializer(data=i)   
            module_serializer.is_valid()
            module_serializer.save()
        print(module_serializer)
        return Response({
            "data": "true",
            "status_code": 200,
            "message": "Get Device data Lists",
            "response": selected_data
        }, status=status.HTTP_200_OK)
    


        
        
# def greenBuilding():
#     while True:

#         try:
            
#             bms_module = BmsGreenBuildingManageDevice.objects.all()        
#             data_user= bms_module.values()
#             a = data_user[0]['model_no']
#             last_token = getToken()
#             # print(last_token)
#             base_api = 'https://dashboard.airveda.com/api/data/last-hour/'
#             headers = {
#                 'Authorization': f'Bearer {last_token}',
#                 'Content-Type': 'application/json'
#             }
#             data = {
#                 "deviceId":a
#             }
#             response = requests.post(base_api, headers=headers, json=data)
#             device_data = response.json()
#             # print(device_data) 
#             leed_data = []
#             leed_data.clear()
#             for item in device_data['data']: 
#                 leed_data1 = {
#                         "leed_data": {
#                             "temperature": item["temperature"],
#                             "pm10": item["pm10"],
#                             "voc": item["voc"],
#                             "humidity": item["humidity"],
#                             "aqi": item["aqi"],
#                             "pm25": item["pm25"],
#                             "viral_index": item["viral_index"]
#                         }
#                     }
#                 # print(leed_data)
#                 module_serializer = BmsGreenBuildingManageDataSerializersss(data=leed_data1)
#                 # print(module_serializer)
#                 if module_serializer.is_valid():
#                     module_serializer.save()
#                     # print("i anm pass")
#                     leed_data.append(module_serializer)
#                 else:
#                     # Handle invalid data, e.g., log the error or skip saving this data.
#                     print("Invalid data:", module_serializer.errors)
#                     leed_data.append(leed_data1)
#         except:
#             pass




def greenBuilding():
    while True:

        try:
            
            bms_module = BmsGreenBuildingManageDevice.objects.all()        
            data_user= bms_module.values()
            a = data_user[0]['model_no']
            last_token = getToken()
            b = data_user[0]['id']
            # print(last_token)
            base_api = 'https://dashboard.airveda.com/api/data/last-hour/'
            headers = {
                'Authorization': f'Bearer {last_token}',
                'Content-Type': 'application/json'
            }
            data = {
                "deviceId":a
            }
            response = requests.post(base_api, headers=headers, json=data)
            device_data = response.json()
            # print(device_data) 
            selected_data = []
            selected_data.clear()
            for item in device_data['data']:
                selected_item = {
                    "manage_device":b,
                    "temperature": item["temperature"],
                    "pm10": item["pm10"],
                    "voc": item["voc"],
                    "humidity": item["humidity"],
                    "aqi": item["aqi"],
                    "pm25": item["pm25"],
                    "viral_index": item["viral_index"]
                }
                selected_data.append(selected_item)
            # aqi = device_data['data']
            # print(aqi)
            for i in selected_data:
                module_serializer = BmsGreenBuildingManageDataSerializer(data=i)   
                module_serializer.is_valid()
                module_serializer.save()
        except:
            pass


def getToken():
        bms_module = BmsGreenBuildingManageDevice.objects.all()     
        data_user= bms_module.values()
        a = data_user[0]['user_name']
        # print(a)
       
        base_api = data_user[0]['api_key']
        data = {
        "email": data_user[0]['user_name'],
        "password": data_user[0]['password']
        }
        
        response = requests.post(str(base_api),json=data)
        data = response.json()
        # print(data)
        token = data['idToken']
        return token
    
    
@api_view(['GET','POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def lastDatassss(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingManageDevice.objects.all()        
        # # print(bms_module.values())
        data_user= bms_module.values()
        # a = data_user[1]['model_no']
        # b = data_user[0]['id']
        # print(a)
        # print(token[0]['idToken'])
        # data = request.data
    
        last_token = "1cd4e151-dd78-49b7-a1fb-2ab856d22584"
        # print(type(last_token))
        
        # device_id = data['device_id']
        base_api = 'https://aura-b2b-rest.web.app/api/v1/org/qWeLCTC8yiQjvNtpRkMiuNzWIX93/devices/32002f000947393133303834/sensors'
    
        
        headers = {
            'Authorization': f'Bearer {last_token}'
            # 'Content-Type': 'Authorization'
        }
        # data = {
        #     "deviceId":a
        # }

        # response = requests.post(base_api, headers=headers,json=data)
        response = requests.get(base_api, headers=headers)
        
        # print('pass')
        
        # print(response)
        # print(response.status_code)
        # print(response.content)
        
        # data_dict = json.loads(response)
        
        device_data = response.json()
        
        # device_data = json.loads(response)
        # print(device_data)
        return Response({
            "data": "true",
            "status_code": 200,
            "message": "Get Device data Lists",
            "response": device_data
        }, status=status.HTTP_200_OK)
        
        
        
        

@api_view(['GET','POST', 'DELETE'])
# @permission_classes([IsAuthenticated])
def lastDeviceDatasss(request):
    if request.method == 'GET':
        bms_module = BmsGreenBuildingManageDevice.objects.all()        
        # print(bms_module.values())
        data_user= bms_module.values()
        # a = data_user[1]['model_no']
        # b = data_user[2]['id']
        # print(token[0]['idToken'])
        # data = request.data
        # idToken= data['idToken']
        # token = ""
        last_token = "1cd4e151-dd78-49b7-a1fb-2ab856d22584"
        # print(type(last_token))
        
        # device_id = data['device_id']
        base_api = 'https://aura-b2b-rest.web.app/api/v1/org/qWeLCTC8yiQjvNtpRkMiuNzWIX93/devices/32002f000947393133303834/sensors'
    
        
        headers = {
            'Authorization': f'Bearer {last_token}'
            # 'Content-Type': 'Authorization'
        }

        response = requests.get(base_api, headers=headers)
        device_data = response.json()
        
        # print(device_data)
        # print(device_data['data']['aqi']) 
        
        selected_data = []
        selected_data.clear()
        
        
        # for item in device_data['data']:
        #     value = device_data['data'][item]['value']
        #     units = device_data['data'][item]['units']
        #     print(f"{item}: {value} {units}")
        
        
        for item in device_data['data']:
            
            # print(item)
            
    #         selected_item = {
    #     "temperature": device_data['data']['temperature']['value']
    # }
            
            selected_item = {
                "manage_device": 2,
                "temperature": device_data['data']['temperature']['value'],
                "pm10": device_data['data']['pm10']['value'],
                "voc": device_data['data']['voc']['value'],
                "humidity": device_data['data']['humidity']['value'],
                "aqi": device_data['data']['aqi']['value'],
                # "pm25":device_data['data']['pm25']['value'],
                # "viral_index": device_data['data']['viral_index']['value'],
            }
            # print(selected_item)
            selected_data.append(selected_item)
            
            print(selected_data)
        
        for i in selected_data:
            module_serializer = BmsGreenBuildingManageDataSerializer(data=i)   
            module_serializer.is_valid()
            module_serializer.save()
            
        # print(module_serializer)
        return Response({
            "data": "true",
            "status_code": 200,
            "message": "Get Device data Lists",
            "response": selected_data
        }, status=status.HTTP_200_OK)
        
        
